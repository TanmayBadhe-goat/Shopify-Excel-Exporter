"""
excel_updater.py — Mode 2: Update an existing Excel workbook with Shopify data.

Reads an existing .xlsx file, detects the Order Number column, fetches each
order from Shopify, then fills in Picture, Color, and Size columns.
"""

import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from pathlib import Path

from .shopify_api import ShopifyAPI
from .image_downloader import ImageDownloader
from .image_resolver import ProductImageResolver
from .data_utils import extract_color_robust, extract_size_robust
from .utils import logger

# Aliases for column detection
_ORDER_NUMBER_ALIASES = {"order number", "order no", "order #", "order no.", "ordernumber", "order_number"}
_PICTURE_ALIASES      = {"picture", "image", "photo", "product image", "img"}
_COLOR_ALIASES        = {"color", "colour", "color name"}
_SIZE_ALIASES         = {"size", "sizes"}

def _detect_columns(sheet) -> dict:
    col_map = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        val = str(cell.value or "").strip().lower()
        if val in _ORDER_NUMBER_ALIASES:
            col_map["order_number"] = col_idx
        elif val in _PICTURE_ALIASES:
            col_map["picture"] = col_idx
        elif val in _COLOR_ALIASES:
            col_map["color"] = col_idx
        elif val in _SIZE_ALIASES:
            col_map["size"] = col_idx
    return col_map

def _has_image_at(ws, row, col):
    col_letter = get_column_letter(col)
    cell_ref = f"{col_letter}{row}"
    for img in ws._images:
        if img.anchor == cell_ref:
            return True
    return False

def update_excel(filepath: str, log_fn=None, progress_fn=None, status_fn=None):
    def _log(msg):
        logger.info(msg)
        if log_fn: log_fn(msg)

    def _progress(val):
        if progress_fn: progress_fn(val)

    def _status(msg):
        if status_fn: status_fn(msg)

    path = Path(filepath)
    _log(f"Opening workbook: {path.name}")
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    col_map = _detect_columns(ws)
    if "order_number" not in col_map:
        raise ValueError("Order Number column not found.")

    on_col = col_map["order_number"]
    pic_col = col_map.get("picture")
    clr_col = col_map.get("color")
    sz_col = col_map.get("size")

    unique_orders = set()
    for row_idx in range(2, ws.max_row + 1):
        raw = str(ws.cell(row=row_idx, column=on_col).value or "").strip()
        order_num = raw.lstrip("#").strip()
        if order_num: unique_orders.add(order_num)

    _log(f"Found {len(unique_orders)} unique orders to fetch.")
    
    api = ShopifyAPI()
    resolver = ProductImageResolver(api)
    downloader = ImageDownloader()
    order_cache = {}

    # Fetch orders
    for idx, order_num in enumerate(sorted(unique_orders), start=1):
        _progress((idx / len(unique_orders)) * 30)
        _status(f"Fetching order {idx}/{len(unique_orders)}")
        try:
            # Using name search with # prefix for more accurate matching
            search_name = f"#{order_num}"
            resp = api._make_request("orders.json", params={"name": search_name, "status": "any", "limit": 10})
            orders_list = resp.json().get("orders", [])
            # Robust matching: check both order_number (int) and name (string)
            matched = next((o for o in orders_list if str(o.get("order_number")) == order_num or o.get("name") == search_name), None)
            order_cache[order_num] = matched
        except Exception as e:
            _log(f"Error fetching order #{order_num}: {e}")
            order_cache[order_num] = None

    # Update rows
    total_rows = ws.max_row - 1
    for row_idx in range(2, ws.max_row + 1):
        _progress(30 + ((row_idx - 1) / total_rows) * 70)
        _status(f"Updating row {row_idx}/{ws.max_row}")
        
        raw_order = str(ws.cell(row=row_idx, column=on_col).value or "").strip()
        order_num = raw_order.lstrip("#").strip()
        if not order_num or not order_cache.get(order_num): continue

        order = order_cache[order_num]
        line_items = order.get("line_items", [])
        if not line_items: continue
        
        # Simple heuristic: if there's only one item, use it. 
        # If there are multiple, we might need more logic, but for now we take the first.
        line_item = line_items[0]
        p_id = line_item.get("product_id")
        v_id = line_item.get("variant_id")

        product = None
        if p_id:
            # Check the resolver's cache first
            product = resolver._product_cache.get(p_id)
            if product is None:
                product = api.get_product(p_id)
                resolver._product_cache[p_id] = product or "__NO_IMAGE__"
            elif product == "__NO_IMAGE__":
                product = None

        # Update Color
        if clr_col and not ws.cell(row=row_idx, column=clr_col).value:
            color_val = extract_color_robust(line_item, product)
            ws.cell(row=row_idx, column=clr_col).value = color_val

        # Update Size
        if sz_col and not ws.cell(row=row_idx, column=sz_col).value:
            size_val = extract_size_robust(line_item, product)
            ws.cell(row=row_idx, column=sz_col).value = size_val

        # Update Picture
        if pic_col and not _has_image_at(ws, row_idx, pic_col):
            url = resolver.get_image_url(line_item)
            if url:
                # Ensure we use a unique file_id for the download cache
                local_path = downloader.download_image(url, f"{p_id}_{v_id}")
                if local_path and Path(local_path).exists():
                    try:
                        img = ExcelImage(str(local_path))
                        # Adjust row height for the image
                        ws.row_dimensions[row_idx].height = 75
                        img.width = 90
                        img.height = 70
                        # Calculate the anchor cell properly
                        cell_ref = f"{get_column_letter(pic_col)}{row_idx}"
                        ws.add_image(img, cell_ref)
                    except Exception as e:
                        _log(f"Error inserting image for row {row_idx}: {e}")

    output_path = path.parent / f"{path.stem}_UPDATED.xlsx"
    wb.save(str(output_path))
    _log(f"Saved updated workbook to: {output_path.name}")
    return output_path
