import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
from pathlib import Path
from utils import logger

class ExcelExporter:
    def __init__(self, export_dir="exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
        # Simplified column configuration
        self.column_configs = [
            ("Image", 15),
            ("Order Number", 15),
            ("Customer Name", 20),
            ("Customer Email", 25),
            ("Phone Number", 15),
            ("Product Name", 35),
            ("Variant Name", 20),
            ("Color", 12),
            ("Size", 10),
            ("Price", 12),
            ("Payment Status", 15)
        ]

    def export_orders_to_excel(self, orders_data, image_paths):
        logger.info(f"Starting Excel export for {len(orders_data)} items...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Shopify Orders"

        # Write headers
        headers = [config[0] for config in self.column_configs]
        sheet.append(headers)

        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        for col_idx, cell in enumerate(sheet[1], start=1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = self.column_configs[col_idx-1][1]

        # Write data rows
        for row_idx, order_item in enumerate(orders_data, start=2):
            row_values = []
            for header, _ in self.column_configs:
                # Map header to data key
                key = header.lower().replace(" ", "_")
                val = order_item.get(key, "")
                row_values.append(val)
            
            # Placeholder for image
            row_values[0] = ""
            sheet.append(row_values)

            # Insert image
            p_id = order_item.get("product_id")
            v_id = order_item.get("variant_id")
            cache_key = (p_id, v_id) if p_id else None
            img_path = image_paths.get(cache_key) if cache_key else None
            
            if img_path and Path(img_path).exists():
                try:
                    img = ExcelImage(img_path)
                    sheet.row_dimensions[row_idx].height = 75
                    img.width = 90
                    img.height = 70
                    sheet.add_image(img, f"A{row_idx}")
                except Exception as e:
                    logger.error(f"Failed to insert image {img_path}: {e}")

            # Cell formatting
            for col_idx, cell in enumerate(sheet[row_idx], start=1):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                
                header_name = headers[col_idx-1]
                if header_name == "Price":
                    try:
                        cell.value = float(cell.value) if cell.value else 0.0
                        cell.number_format = '"$"#,##0.00'
                    except:
                        pass

        # Freeze first row and add auto-filter
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        # Save file
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = self.export_dir / f"Shopify_Orders_{timestamp}.xlsx"
        workbook.save(filename)
        logger.info(f"Excel export complete: {filename}")
        return filename
