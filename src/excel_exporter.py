import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
from pathlib import Path
from .utils import logger

class ExcelExporter:
    def __init__(self, export_dir="exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
        # Base column configuration
        self.base_column_configs = [
            ("Image", 15),
            ("Order Number", 15),
            ("Customer Name", 20),
            ("Customer Email", 25),
            ("Phone Number", 15),
            ("Product Name", 35),
            ("Variant Name", 20),
            ("Color", 12),
            ("Size", 10),
            ("Price", 12),
            ("Quantity", 10),
            ("Payment Status", 15),
        ]

    def export_orders_to_excel(self, orders_data, image_paths, remittance_data=None, is_remittance_report=False, include_images=True, enabled_columns=None):
        """
        Export orders to Excel. 
        - include_images: If False, skips image column and embedding.
        - remittance_data: If provided, adds remittance columns and highlights matches.
        - enabled_columns: Optional list of column names to include.
                           If None, uses all default columns (backwards compatible).
        """
        logger.info(f"Starting Excel export for {len(orders_data)} items...")
        logger.info(f"Include Product Images: {'YES' if include_images else 'NO'}")
        logger.info(f"Custom columns: {'Yes' if enabled_columns else 'No (default)'}")
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Shopify Orders"

        # Determine columns
        column_configs = []
        for config in self.base_column_configs:
            if config[0] == "Image" and not include_images:
                continue
            # Apply column filtering if enabled_columns is specified
            if enabled_columns is not None and config[0] not in enabled_columns:
                continue
            column_configs.append(config)
            
        if is_remittance_report:
            column_configs.extend([
                ("Remittance Amount", 18),
                ("Remittance Date", 18),
                ("Remittance Status", 18)
            ])

        # Write headers
        headers = [config[0] for config in column_configs]
        sheet.append(headers)

        # Formatting styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        remittance_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid") # Light Green

        # Format headers
        for col_idx, cell in enumerate(sheet[1], start=1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = column_configs[col_idx-1][1]

        matched_count = 0

        # Write data rows
        for row_idx, order_item in enumerate(orders_data, start=2):
            row_values = []
            
            # 1. Map base fields
            for header, _ in self.base_column_configs:
                if header == "Image":
                    if include_images:
                        row_values.append("") # Placeholder for image
                    continue
                key = header.lower().replace(" ", "_")
                val = order_item.get(key, "")
                row_values.append(val)
            
            # 2. Map remittance fields if applicable
            is_matched = False
            if is_remittance_report:
                raw_order = str(order_item.get("order_number") or "").lstrip("#").strip()
                match = remittance_data.get(raw_order) if remittance_data else None
                
                if match:
                    is_matched = True
                    matched_count += 1
                    row_values.append(match.get("amount", ""))
                    row_values.append(match.get("date", ""))
                    row_values.append("Paid")
                    logger.info(f"Matched Order #{raw_order}")
                else:
                    row_values.extend(["", "", ""])
            
            sheet.append(row_values)

            # 3. Insert image if enabled
            if include_images:
                p_id = order_item.get("product_id")
                v_id = order_item.get("variant_id")
                cache_key = (p_id, v_id) if p_id else None
                img_path = image_paths.get(cache_key) if cache_key else None
                
                if img_path and Path(img_path).exists():
                    try:
                        img = ExcelImage(img_path)
                        sheet.row_dimensions[row_idx].height = 75
                        img.width = 90
                        img.height = 70
                        sheet.add_image(img, f"A{row_idx}")
                    except Exception as e:
                        logger.error(f"Failed to insert image {img_path}: {e}")

            # 4. Row Formatting
            for col_idx, cell in enumerate(sheet[row_idx], start=1):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                
                # Apply green fill if matched
                if is_matched:
                    cell.fill = remittance_fill
                
                # Format Price columns
                header_name = headers[col_idx-1]
                if "Price" in header_name or "Amount" in header_name:
                    try:
                        cell.value = float(cell.value) if cell.value else 0.0
                        cell.number_format = '"$"#,##0.00'
                    except:
                        pass

        # Final adjustments
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        # Save file
        timestamp = datetime.now().strftime("%Y-%m-%d")
        if is_remittance_report:
            filename = self.export_dir / f"Orders_Remittance_{timestamp}_{datetime.now().strftime('%H%M%S')}.xlsx"
            logger.info(f"Matched: {matched_count}")
            logger.info(f"Not Matched: {len(orders_data) - matched_count}")
        else:
            filename = self.export_dir / f"Shopify_Orders_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
            
        workbook.save(filename)
        logger.info(f"Export Complete: {filename}")
        return filename
